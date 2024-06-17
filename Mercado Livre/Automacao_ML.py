from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pandas as pd

valores_fixos = {
    '\u200bCondição': 'Novo',
    'Forma de anunciar': 'Lista geral',
    'Código universal de produto': 'O produto não tem código cadastrado',
    'Forma de envio': 'Mercado Envios',
    'Frete': 'Por conta do comprador',
    'Retirar pessoalmente': 'Concordo',
    'Tipo de garantia': 'Garantia do vendedor',
    'Tempo de garantia': 30,
    'Unidade de Tempo de garantia': 'dias',
    'Tipo de produto': 'N/A'
}

df_valores_fixos = pd.DataFrame(valores_fixos, index=[0])

# Função para ler as colunas da linha 3 da planilha
def ler_colunas(sheet):
    colunas = {}
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=3, column=col).value
        if header:
            colunas[header.lower()] = col
    return colunas

# Função para mapear colunas do arquivo de entrada para colunas do arquivo de anúncio
def mapear_colunas(df_entrada, sheet_anuncio, df_valores_fixos):
    mapeamento = {}
    colunas_anuncio = [cell.value.strip() for cell in sheet_anuncio[3]]

    # Map columns from df_valores_fixos
    for col_nome in df_valores_fixos.columns:
        if col_nome.strip() in colunas_anuncio:
            col_index = colunas_anuncio.index(col_nome.strip()) + 1
            mapeamento[col_nome.strip()] = col_index

    # Map columns from df_entrada
    for col_nome in df_entrada.columns:
        if col_nome.strip() in colunas_anuncio:
            col_index = colunas_anuncio.index(col_nome.strip()) + 1
            mapeamento[col_nome.strip()] = col_index

    return mapeamento

# Função para configurar a célula
def configurar_celula(sheet, row, col, value, alignment=False):
    if isinstance(col, str):
        for cell in sheet[3]:  # Alterei para a linha 3, onde os cabeçalhos estão
            if cell.value == col:
                col = cell.column
                break
        else:
            print(f"Erro: Não foi possível encontrar a coluna '{col}'")
            return
    cell = sheet.cell(row=row, column=col)
    cell.value = value
    if alignment:
        cell.alignment = Alignment(wrap_text=True)

# Função para calcular o valor final
def calcular_valor_final(valor_receber, porcentagem, taxa_fixa, taxa_frete):
    if valor_receber <= (79 - taxa_fixa) * (1 - porcentagem):
        valor_final = (valor_receber + taxa_fixa) / (1 - porcentagem)
    else:
        valor_final = (valor_receber + taxa_frete + 3) / (1 - porcentagem)
    return valor_final

# Função para alterar linha
def alterar_linha(sheet, linha_inicio, qtd, df_entrada, df_valores, idx, mapeamento_colunas, porcentagem_classico, porcentagem_premium, taxa_fixa, taxa_frete):
    for x in range(linha_inicio, linha_inicio + qtd):

        for col_nome, col_index in mapeamento_colunas.items():
            if col_nome in df_valores.columns:
                value = df_valores.at[0, col_nome]
                if pd.isna(value):
                    value = 'N/A'
            elif col_nome in df_entrada.columns:
                value = df_entrada.iloc[idx][col_nome]
                if pd.isna(value):
                    value = 'N/A'
            else:
                print(f"Column '{col_nome}' not found in df_valores or df_entrada.")
                value = 'N/A'

            configurar_celula(sheet, x, col_index, value)

        # Preencher colunas que não foram preenchidas ainda com 'N/A'
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=x, column=col).value is None:
                configurar_celula(sheet, x, col, 'N/A')

        valor_receber = df_entrada.at[idx, 'valor']

        # Specify the content you want to find
        content_to_find_preco = 'Preço [R$]'
        content_to_find_tipo = 'Tipo de anúncio'

        # Initialize the column indices
        preco_column_index = int
        tipo_column_index = int

        # Iterate through the cells of the header row (assuming it's the first row)
        for cell in sheet[3]:
            if cell.value == content_to_find_preco:
                preco_column_index = cell.column
            elif cell.value == content_to_find_tipo:
                tipo_column_index = cell.column

        if x % 2 == 0:
            configurar_celula(sheet, x, tipo_column_index, "Clássico")

            valor_final = calcular_valor_final(valor_receber, porcentagem_classico, taxa_fixa, taxa_frete)
            configurar_celula(sheet, x, preco_column_index, valor_final)
        else:
            configurar_celula(sheet, x, tipo_column_index, "Premium")

            valor_final = calcular_valor_final(valor_receber, porcentagem_premium, taxa_fixa, taxa_frete)
            configurar_celula(sheet, x, preco_column_index, valor_final)

# Função para processar entrada de kits
def processar_kits():
    kits = {}
    entrada_kits = input("Digite os kits no formato 'nome:quantidade' separados por vírgula (ex: 'Kit 2 Unid:2, Kit 4 Unid:4'): ")
    for kit in entrada_kits.split(','):
        nome_kit, quantidade_kit = kit.split(':')
        kits[nome_kit.strip()] = int(quantidade_kit.strip())
    return kits

Arquivo_nm = input("Digite o nome da planilha que o ML gerou (Case sensitive): ")

# Importar planilhas
df_entrada = pd.read_excel('Inputs/Automação_input.xlsx')

# Carrega o arquivo Excel de saída
workbook = load_workbook(f'Inputs/{Arquivo_nm}')

# Função para carregar ou criar uma planilha baseada na categoria
def carregar_planilha(workbook, categoria):
    return workbook[f'{categoria}']

# Função para carregar a planilha de categorias e valores
def carregar_planilha_categorias(caminho_arquivo):
    return pd.read_excel(caminho_arquivo)

# Carrega a planilha de categorias
df_categorias = carregar_planilha_categorias('Inputs/Categorias.xlsx')

# Loop principal
while True:
    linha_atual = {}
    for index, row in df_entrada.iterrows():

        categoria = row['Categoria']
        categoria_info = df_categorias[df_categorias['Categoria'] == categoria].iloc[0]
        porcentagem_classico = categoria_info['Porcentagem_Classico'] / 100
        porcentagem_premium = categoria_info['Porcentagem_Premium'] / 100
        taxa_fixa = categoria_info['Preço_Fixo']
        taxa_frete = categoria_info['Preço_Frete']
        print(taxa_fixa)

        if categoria not in linha_atual:
            sheet = carregar_planilha(workbook, categoria)
            colunas_planilha = ler_colunas(sheet)
            mapeamento_colunas = mapear_colunas(df_entrada, sheet, df_valores_fixos)
            linha_atual[categoria] = 8

        tem_kits = input(f"\nHá kits para o produto {row['Título: informe o produto, marca, modelo e destaque as características principais']} (s/n)? ").lower() == 's'
        if tem_kits:
            kits = processar_kits()

            sku = df_entrada.at[index, 'SKU']
            df_entrada.at[index, 'SKU'] = f"{sku}-1"
            alterar_linha(sheet, linha_atual[categoria], 2, df_entrada, df_valores_fixos, index, mapeamento_colunas, porcentagem_classico, porcentagem_premium, taxa_fixa, taxa_frete)
            linha_atual[categoria] += 2

            df_entrada.at[index, 'SKU'] = f"{sku}"

            nome_pai = df_entrada.at[index, 'Título: informe o produto, marca, modelo e destaque as características principais']
            sku_pai = df_entrada.at[index, 'SKU']
            valor_pai = df_entrada.at[index, 'valor']

            for nome_kit, qtde in kits.items():
                sku = sku_pai
                nome = nome_pai
                valor = valor_pai

                df_entrada.at[index, 'Título: informe o produto, marca, modelo e destaque as características principais'] = f"{nome_kit} {nome}"
                df_entrada.at[index, 'SKU'] = f"{sku}-{qtde}"
                df_entrada.at[index, 'valor'] = valor * qtde
                alterar_linha(sheet, linha_atual[categoria], 2, df_entrada, df_valores_fixos, index, mapeamento_colunas, porcentagem_classico, porcentagem_premium, taxa_fixa, taxa_frete)
                linha_atual[categoria] += 2
        else:
            sku = df_entrada.at[index, 'SKU']
            df_entrada.at[index, 'SKU'] = f"{sku}-1"
            alterar_linha(sheet, linha_atual[categoria], 2, df_entrada, df_valores_fixos, index, mapeamento_colunas, porcentagem_classico, porcentagem_premium, taxa_fixa, taxa_frete)
            linha_atual[categoria] += 2

    repetir = input('Deseja repetir (s/n)?').lower() == 's'
    if not repetir:
        break

# Salva as alterações no arquivo
workbook.save(f'Output/{Arquivo_nm}')
